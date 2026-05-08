from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.core.cache import cache
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from .models import Materia, Inscripcion, Actividad, Calificacion, Asistencia, HorarioCurso, PeriodoHorario, ClaseHorario
from .serializers import (
    MateriaSerializer, InscripcionSerializer, ActividadSerializer,
    CalificacionSerializer, AsistenciaSerializer, AsistenciaBulkItemSerializer,
    HorarioCursoSerializer,
)
from accounts.models import Estudiante, Docente
from accounts.serializers import EstudianteSerializer


def _solo_administrativo(request):
    if request.user.role != 'administrativo':
        return Response({'error': 'Solo el personal administrativo puede realizar esta acción.'}, status=status.HTTP_403_FORBIDDEN)
    return None


# ── Materias ──────────────────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def materias_list(request):
    if request.method == 'GET':
        cache_key = 'materias_list_all'
        data = cache.get(cache_key)
        if not data:
            materias = Materia.objects.filter(estado=True).select_related('curso', 'docente__user').order_by('id')
            data = MateriaSerializer(materias[:500], many=True).data
            cache.set(cache_key, data, 300)  # Cache for 5 minutes
        return Response(data)
    serializer = MateriaSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def materia_detail(request, pk):
    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
    if request.method == 'GET':
        return Response(MateriaSerializer(materia).data)
    if request.method == 'PUT':
        serializer = MateriaSerializer(materia, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    materia.estado = False
    materia.save()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def materia_estudiantes(request, pk):
    """Todos los estudiantes activos del curso al que pertenece la materia."""
    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
    estudiantes = Estudiante.objects.filter(curso=materia.curso, estado='activo').select_related('user', 'curso')
    return Response(EstudianteSerializer(estudiantes, many=True).data)


# ── Inscripciones (solo administrativo) ───────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def inscripciones_list(request):
    err = _solo_administrativo(request)
    if err:
        return err
    if request.method == 'GET':
        qs = Inscripcion.objects.select_related('estudiante__user', 'estudiante__curso', 'curso').all().order_by('-id')
        estudiante_id = request.query_params.get('estudiante')
        curso_id = request.query_params.get('curso')
        estado = request.query_params.get('estado')
        if estudiante_id:
            qs = qs.filter(estudiante_id=estudiante_id)
        if curso_id:
            qs = qs.filter(curso_id=curso_id)
        if estado:
            qs = qs.filter(estado=estado)
        
        paginator = PageNumberPagination()
        paginator.page_size = 50
        result_page = paginator.paginate_queryset(qs, request)
        serializer = InscripcionSerializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)

    serializer = InscripcionSerializer(data=request.data)
    if serializer.is_valid():
        from accounts.models import Estudiante as EstudianteModel, Curso as CursoModel
        try:
            estudiante = EstudianteModel.objects.get(pk=serializer.validated_data['estudiante_id'])
            curso = CursoModel.objects.get(pk=serializer.validated_data['curso_id'])
        except (EstudianteModel.DoesNotExist, CursoModel.DoesNotExist):
            return Response({'error': 'Estudiante o curso no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        inscripcion = Inscripcion(
            estudiante=estudiante,
            curso=curso,
            estado=serializer.validated_data.get('estado', 'activo'),
            registrada_por=request.user,
        )
        inscripcion.save()
        return Response(InscripcionSerializer(inscripcion).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def inscripcion_detail(request, pk):
    err = _solo_administrativo(request)
    if err:
        return err
    try:
        inscripcion = Inscripcion.objects.select_related('estudiante__user', 'curso').get(pk=pk)
    except Inscripcion.DoesNotExist:
        return Response({'error': 'Inscripción no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
    if request.method == 'GET':
        return Response(InscripcionSerializer(inscripcion).data)
    if request.method == 'PUT':
        allowed = {k: request.data[k] for k in ('estado',) if k in request.data}
        if not allowed:
            return Response({'error': 'Solo se puede actualizar el estado.'}, status=status.HTTP_400_BAD_REQUEST)
        inscripcion.estado = allowed.get('estado', inscripcion.estado)
        inscripcion.save()
        return Response(InscripcionSerializer(inscripcion).data)
    inscripcion.estado = 'retirado'
    inscripcion.save()
    inscripcion.estudiante.curso = None
    inscripcion.estudiante.save(update_fields=['curso'])
    return Response(status=status.HTTP_204_NO_CONTENT)



# ── Actividades ────────────────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def materia_actividades(request, pk):
    """Actividades de una materia específica."""
    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        trimestre = request.query_params.get('trimestre')
        actividades = Actividad.objects.filter(materia=materia, es_plantilla=False).order_by('trimestre', 'fecha', 'created_at')
        if trimestre:
            actividades = actividades.filter(trimestre=trimestre)
        return Response(ActividadSerializer(actividades, many=True).data)

    data = request.data.copy()
    data['materia'] = pk
    serializer = ActividadSerializer(data=data)
    if serializer.is_valid():
        serializer.save(creada_por=request.user, materia=materia)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def actividad_detail(request, pk):
    try:
        actividad = Actividad.objects.get(pk=pk)
    except Actividad.DoesNotExist:
        return Response({'error': 'Actividad no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'PUT':
        serializer = ActividadSerializer(actividad, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    actividad.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def guardar_plantilla(request, pk, pk2):
    """Duplica una actividad de una materia como plantilla."""
    try:
        actividad = Actividad.objects.get(pk=pk2)
    except Actividad.DoesNotExist:
        return Response({'error': 'Actividad no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
    plantilla = Actividad.objects.create(
        materia=None,
        nombre=actividad.nombre,
        descripcion=actividad.descripcion,
        tipo=actividad.tipo,
        ponderacion=actividad.ponderacion,
        es_plantilla=True,
        creada_por=request.user,
    )
    return Response(ActividadSerializer(plantilla).data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mis_plantillas(request):
    """Plantillas del docente autenticado."""
    plantillas = Actividad.objects.filter(creada_por=request.user, es_plantilla=True).order_by('nombre')
    return Response(ActividadSerializer(plantillas, many=True).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def usar_plantilla(request, pk, plantilla_pk):
    """Crea una actividad en la materia basada en una plantilla."""
    try:
        materia = Materia.objects.get(pk=pk)
        plantilla = Actividad.objects.get(pk=plantilla_pk, es_plantilla=True)
    except (Materia.DoesNotExist, Actividad.DoesNotExist):
        return Response({'error': 'Recurso no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    actividad = Actividad.objects.create(
        materia=materia,
        nombre=plantilla.nombre,
        descripcion=plantilla.descripcion,
        tipo=plantilla.tipo,
        ponderacion=plantilla.ponderacion,
        es_plantilla=False,
        creada_por=request.user,
    )
    return Response(ActividadSerializer(actividad).data, status=status.HTTP_201_CREATED)


# ── Calificaciones (centro de notas) ──────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def centro_notas(request, pk):
    """Devuelve el grid de notas: {actividades, filas de estudiantes con notas}."""
    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    trimestre = request.query_params.get('trimestre')
    actividades = Actividad.objects.filter(materia=materia, es_plantilla=False).order_by('trimestre', 'fecha', 'created_at')
    if trimestre:
        actividades = actividades.filter(trimestre=trimestre)
    estudiantes = Estudiante.objects.filter(curso=materia.curso, estado='activo').select_related('user')

    if request.user.role == 'estudiante':
        try:
            est_obj = Estudiante.objects.get(user=request.user)
            estudiantes = estudiantes.filter(pk=est_obj.pk)
        except Estudiante.DoesNotExist:
            estudiantes = Estudiante.objects.none()

    calificaciones = Calificacion.objects.filter(actividad__in=actividades).values('estudiante_id', 'actividad_id', 'nota')
    cal_map = {}
    for c in calificaciones:
        if c['estudiante_id'] not in cal_map:
            cal_map[c['estudiante_id']] = {}
        cal_map[c['estudiante_id']][str(c['actividad_id'])] = c['nota']

    actividades_data = ActividadSerializer(actividades, many=True).data
    
    filas = []
    for e in estudiantes:
        filas.append({
            'estudiante_id': e.id,
            'nombre': f"{e.user.first_name} {e.user.last_name}",
            'notas': cal_map.get(e.id, {}),
        })

    return Response({
        'actividades': actividades_data,
        'filas': filas,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def guardar_nota(request):
    """Guarda o actualiza una nota puntual: {estudiante, actividad, nota}."""
    estudiante_id = request.data.get('estudiante')
    actividad_id = request.data.get('actividad')
    nota = request.data.get('nota')

    try:
        estudiante = Estudiante.objects.get(pk=estudiante_id)
        actividad = Actividad.objects.get(pk=actividad_id)
    except (Estudiante.DoesNotExist, Actividad.DoesNotExist):
        return Response({'error': 'Recurso no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    obj, _ = Calificacion.objects.update_or_create(
        estudiante=estudiante,
        actividad=actividad,
        defaults={'nota': nota, 'creada_por': request.user},
    )
    return Response(CalificacionSerializer(obj).data)


# ── Asistencia ─────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def asistencia_materia(request, pk):
    """Lista asistencias de una materia, opcionalmente filtradas por fecha."""
    fecha = request.query_params.get('fecha')
    qs = Asistencia.objects.filter(materia_id=pk).select_related('estudiante__user')
    if fecha:
        qs = qs.filter(fecha=fecha)
    return Response(AsistenciaSerializer(qs, many=True).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def registrar_asistencia_bulk(request, pk):
    """Registra o actualiza la asistencia de varios estudiantes para una fecha.
    Body: { fecha: 'YYYY-MM-DD', registros: [{estudiante, estado, motivo}] }
    """
    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    from .models import Asistencia as AsistenciaModel
    
    fecha = request.data.get('fecha')
    registros = request.data.get('registros', [])

    if not fecha:
        return Response({'error': 'Se requiere la fecha.'}, status=status.HTTP_400_BAD_REQUEST)

    serializer = AsistenciaBulkItemSerializer(data=registros, many=True)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Optimizamos usando bulk_create con update_conflicts (PostgreSQL 9.5+)
    nuevas_asistencias = []
    for r in serializer.validated_data:
        nuevas_asistencias.append(AsistenciaModel(
            estudiante_id=r['estudiante'],
            materia=materia,
            fecha=fecha,
            estado=r['estado'],
            motivo=r.get('motivo', ''),
            registrada_por=request.user
        ))

    AsistenciaModel.objects.bulk_create(
        nuevas_asistencias,
        update_conflicts=True,
        unique_fields=['estudiante', 'materia', 'fecha'],
        update_fields=['estado', 'motivo', 'registrada_por']
    )

    # Devolvemos el resultado actualizado
    qs = AsistenciaModel.objects.filter(materia=materia, fecha=fecha).select_related('estudiante__user')
    return Response(AsistenciaSerializer(qs, many=True).data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def resumen_asistencia(request, pk):
    """Resumen de faltas y licencias por estudiante en una materia."""
    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    estudiantes = Estudiante.objects.filter(curso=materia.curso, estado='activo').select_related('user')

    if request.user.role == 'estudiante':
        try:
            est_obj = Estudiante.objects.get(user=request.user)
            estudiantes = estudiantes.filter(pk=est_obj.pk)
        except Estudiante.DoesNotExist:
            estudiantes = Estudiante.objects.none()

    from django.db.models import Count, Q
    
    resultado = []
    # Obtenemos estadísticas de una sola vez para todos los estudiantes
    stats = Asistencia.objects.filter(
        materia=materia, 
        estudiante__in=estudiantes
    ).values('estudiante_id').annotate(
        presentes=Count('id', filter=Q(estado='P')),
        faltas=Count('id', filter=Q(estado='F')),
        licencias=Count('id', filter=Q(estado='L')),
    )
    
    # Mapeamos para acceso rápido
    stats_map = {s['estudiante_id']: s for s in stats}

    for e in estudiantes:
        s = stats_map.get(e.id, {'presentes': 0, 'faltas': 0, 'licencias': 0})
        resultado.append({
            'estudiante_id': e.id,
            'nombre': f"{e.user.first_name} {e.user.last_name}",
            'presentes': s['presentes'],
            'faltas': s['faltas'],
            'licencias': s['licencias'],
        })
    return Response(resultado)


# ── Última asistencia ──────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ultima_asistencia(request, pk):
    """La sesión de asistencia más reciente de la materia antes de una fecha dada."""
    from datetime import date as date_type
    antes_de = request.query_params.get('antes_de', str(date_type.today()))
    ultima_fecha = (
        Asistencia.objects.filter(materia_id=pk, fecha__lt=antes_de)
        .order_by('-fecha')
        .values_list('fecha', flat=True)
        .first()
    )
    if ultima_fecha is None:
        return Response({'fecha': None, 'registros': {}})
    registros = Asistencia.objects.filter(materia_id=pk, fecha=ultima_fecha)
    return Response({
        'fecha': str(ultima_fecha),
        'registros': {str(r.estudiante_id): r.estado for r in registros},
    })


# ── Horario ────────────────────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def horario_curso(request, curso_pk):
    """GET: devuelve el horario del curso. POST: reemplaza el horario completo."""
    from accounts.models import Curso as CursoModel
    from datetime import datetime
    try:
        curso = CursoModel.objects.get(pk=curso_pk)
    except CursoModel.DoesNotExist:
        return Response({'error': 'Curso no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    horario, _ = HorarioCurso.objects.get_or_create(curso=curso)

    if request.method == 'GET':
        horario = HorarioCurso.objects.prefetch_related('periodos__clases__materia').get(pk=horario.pk)
        return Response(HorarioCursoSerializer(horario).data)

    periodos_data = request.data.get('periodos', [])
    horario.periodos.all().delete()

    for p_data in periodos_data:
        periodo = PeriodoHorario.objects.create(
            horario=horario,
            orden=p_data['orden'],
            hora_inicio=p_data['hora_inicio'],
            hora_fin=p_data['hora_fin'],
        )
        for c_data in p_data.get('clases', []):
            materia_id = c_data.get('materia_id') or None
            materia = None
            if materia_id:
                try:
                    materia = Materia.objects.get(pk=materia_id, curso=curso)
                except Materia.DoesNotExist:
                    pass
            ClaseHorario.objects.create(periodo=periodo, dia=c_data['dia'], materia=materia)

    # Auto-update numero_horas for each materia in this curso
    materias_curso = Materia.objects.filter(curso=curso, estado=True)
    for mat in materias_curso:
        total_min = 0
        for p in horario.periodos.prefetch_related('clases').all():
            for clase in p.clases.filter(materia=mat):
                inicio = datetime.combine(datetime.today(), p.hora_inicio)
                fin = datetime.combine(datetime.today(), p.hora_fin)
                total_min += max(0, int((fin - inicio).total_seconds() // 60))
        horas = round(total_min / 60)
        if horas > 0:
            mat.numero_horas = horas
            mat.save(update_fields=['numero_horas'])

    horario.save()
    return Response(HorarioCursoSerializer(horario).data)


# ── Exportación Excel ──────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_asistencia(request, pk):
    """Exporta asistencia de una materia a Excel. Params: fecha_inicio, fecha_fin."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    fecha_inicio = request.query_params.get('fecha_inicio')
    fecha_fin = request.query_params.get('fecha_fin')

    qs = Asistencia.objects.filter(materia=materia).select_related('estudiante__user').order_by('fecha')
    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)

    fechas = sorted(set(str(a.fecha) for a in qs))
    estudiantes = list(
        Estudiante.objects.filter(curso=materia.curso, estado='activo')
        .select_related('user').order_by('user__last_name', 'user__first_name')
    )

    asis_map = {(a.estudiante_id, str(a.fecha)): a.estado for a in qs}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asistencia'

    HDR = PatternFill('solid', fgColor='1565C0')
    HDR_FONT = Font(bold=True, color='FFFFFF')
    CENTER = Alignment(horizontal='center', vertical='center')
    FILLS = {
        'P': PatternFill('solid', fgColor='C8E6C9'),
        'F': PatternFill('solid', fgColor='FFCDD2'),
        'L': PatternFill('solid', fgColor='BBDEFB'),
    }
    TOTALS_FILL = {
        'P': PatternFill('solid', fgColor='2E7D32'),
        'F': PatternFill('solid', fgColor='C62828'),
        'L': PatternFill('solid', fgColor='1565C0'),
    }

    # ── Fila de título ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(fechas) + 5, 3))
    title_cell = ws.cell(row=1, column=1,
        value=f"Asistencia – {materia.nombre} ({materia.codigo})"
              + (f"  {fecha_inicio} al {fecha_fin}" if fecha_inicio or fecha_fin else ""))
    title_cell.font = Font(bold=True, size=13, color='1565C0')
    ws.row_dimensions[1].height = 20

    # ── Cabecera ──
    for col, val in [(1, 'N°'), (2, 'Apellidos y Nombre')]:
        c = ws.cell(row=2, column=col, value=val)
        c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
    for i, fecha in enumerate(fechas):
        c = ws.cell(row=2, column=i + 3, value=fecha)
        c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(i + 3)].width = 13
    for j, lbl in enumerate(['P', 'F', 'L']):
        col = len(fechas) + 3 + j
        c = ws.cell(row=2, column=col, value=lbl)
        c.font = HDR_FONT; c.fill = TOTALS_FILL[lbl]; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 6

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32

    # ── Filas de datos ──
    for i, est in enumerate(estudiantes):
        row = i + 3
        ws.cell(row=row, column=1, value=i + 1).alignment = CENTER
        ws.cell(row=row, column=2, value=f"{est.user.last_name} {est.user.first_name}")
        p = f = l = 0
        for j, fecha in enumerate(fechas):
            estado = asis_map.get((est.id, fecha), '')
            c = ws.cell(row=row, column=j + 3, value=estado)
            c.alignment = CENTER
            if estado in FILLS:
                c.fill = FILLS[estado]
                if estado == 'P': p += 1
                elif estado == 'F': f += 1
                else: l += 1
        for j, val in enumerate([p, f, l]):
            c = ws.cell(row=row, column=len(fechas) + 3 + j, value=val)
            c.alignment = CENTER

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    slug = materia.codigo.replace('/', '-')
    rango = f"_{fecha_inicio}_a_{fecha_fin}" if (fecha_inicio or fecha_fin) else ""
    response['Content-Disposition'] = f'attachment; filename="asistencia_{slug}{rango}.xlsx"'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_notas(request, pk):
    """Exporta notas de una materia a Excel. Param: trimestre (T1/T2/T3)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    try:
        materia = Materia.objects.get(pk=pk)
    except Materia.DoesNotExist:
        return Response({'error': 'Materia no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    trimestre = request.query_params.get('trimestre')

    actividades = Actividad.objects.filter(materia=materia, es_plantilla=False).order_by('trimestre', 'fecha', 'created_at')
    if trimestre:
        actividades = actividades.filter(trimestre=trimestre)
    actividades = list(actividades)

    estudiantes = list(
        Estudiante.objects.filter(curso=materia.curso, estado='activo')
        .select_related('user').order_by('user__last_name', 'user__first_name')
    )

    cal_map = {
        (c[0], c[1]): c[2]
        for c in Calificacion.objects.filter(actividad__in=actividades).values_list('estudiante_id', 'actividad_id', 'nota')
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notas {trimestre or 'General'}"

    HDR = PatternFill('solid', fgColor='1565C0')
    HDR_FONT = Font(bold=True, color='FFFFFF')
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    PROM_HDR = PatternFill('solid', fgColor='E65100')

    def nota_fill(nota):
        if nota is None:
            return None
        if nota >= 65:
            return PatternFill('solid', fgColor='C8E6C9')
        if nota >= 51:
            return PatternFill('solid', fgColor='FFF9C4')
        return PatternFill('solid', fgColor='FFCDD2')

    # ── Fila de título ──
    n_cols = len(actividades) + 4
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 3))
    t = ws.cell(row=1, column=1,
        value=f"Notas – {materia.nombre} ({materia.codigo})"
              + (f"  {trimestre}" if trimestre else "  Todos los trimestres"))
    t.font = Font(bold=True, size=13, color='1565C0')
    ws.row_dimensions[1].height = 20

    # ── Cabecera ──
    for col, val in [(1, 'N°'), (2, 'Apellidos y Nombre')]:
        c = ws.cell(row=2, column=col, value=val)
        c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
    for i, act in enumerate(actividades):
        col = i + 3
        lbl = f"{act.nombre}\n{act.tipo.upper()} · {act.trimestre}"
        c = ws.cell(row=2, column=col, value=lbl)
        c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.row_dimensions[2].height = 36
    prom_col = len(actividades) + 3
    c = ws.cell(row=2, column=prom_col, value='Promedio')
    c.font = HDR_FONT; c.fill = PROM_HDR; c.alignment = CENTER
    ws.column_dimensions[get_column_letter(prom_col)].width = 12
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32

    # ── Filas de datos ──
    for i, est in enumerate(estudiantes):
        row = i + 3
        ws.cell(row=row, column=1, value=i + 1).alignment = CENTER
        ws.cell(row=row, column=2, value=f"{est.user.last_name} {est.user.first_name}")
        notas = []
        for j, act in enumerate(actividades):
            col = j + 3
            nota = cal_map.get((est.id, act.id))
            c = ws.cell(row=row, column=col, value=nota if nota is not None else '')
            c.alignment = CENTER
            if nota is not None:
                notas.append(nota)
                fill = nota_fill(nota)
                if fill:
                    c.fill = fill
        prom_cell = ws.cell(row=row, column=prom_col)
        if notas:
            prom = round(sum(notas) / len(notas), 1)
            prom_cell.value = prom
            fill = nota_fill(prom)
            if fill:
                prom_cell.fill = fill
        prom_cell.alignment = CENTER

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    slug = materia.codigo.replace('/', '-')
    suf = f"_{trimestre}" if trimestre else ""
    response['Content-Disposition'] = f'attachment; filename="notas_{slug}{suf}.xlsx"'
    wb.save(response)
    return response
